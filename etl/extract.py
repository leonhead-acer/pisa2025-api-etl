# import dependent modules
import pandas as pd
import json
import psycopg2
import psycopg2.extras
from psycopg2.extras import RealDictCursor
from pathlib import Path
import re
import numpy as np
import datetime
import glob
from openpyxl import load_workbook

# extract data
def extract_data(filepath: object, select_cols: list, rename_cols: dict) -> object:
    """
       Simple Extract Function in Python with Error Handling
       :param filepath: str, file path to CSV data
       :output: pandas dataframe, extracted from CSV data
    """
    try:
        # Read the CSV file and store it in a dataframe
        df = pd.read_csv(filepath)
        df = df[select_cols]
        df = df.rename(columns={rename_cols})

    # Handle exception if any of the files are missing
    except FileNotFoundError as e:
        print(f"Error: {e}")

    # Handle any other exceptions
    except Exception as e:
        print(f"Error: {e}")

    return df

def postgresql_conn(**kwargs):

    params = kwargs['params']

    con = psycopg2.connect(
        f"dbname={params['database']} user={params['username']} host={params['host']} port = {params['port']} password = {params['password']}",
        cursor_factory=RealDictCursor
    )

    return con

def extract_json(domain: str, nc_dat: object, overwrite: bool, con):

    cur = con.cursor()

    print("Connection to DB established, searching for new records...")

    domain = 'FLA'
    if(domain == 'SCI'):
        regex_dom = '^(S\\d|SA1).*'
    elif(domain == 'FLA'):
        regex_dom = '^FLA\\-.*'

    cnt_name = nc_dat['isoalpha3']
    cnt_code = nc_dat['isocntcd']
    json_file_path = Path(f"./data/db/{domain.lower()}/{domain}_{cnt_name}.json")
    json_file_exists = Path(json_file_path).is_file()
    make_file = (overwrite or not json_file_exists)

    if make_file:

        # fetch_query = fetch_query_create(domain,cnt_code)
        fetch_query = f"""
            SELECT
                row_id, delivery_execution_id, delivery_id, last_update_date, login,
                test_qti_id, test_qti_label, test_qti_title, raw_data,
                raw_data->'metadata'->>'PISA25 Languages' as language
            FROM oat.delivery_results
            WHERE test_qti_id ~ '{regex_dom}'
            AND login ~ '(?<=^.{{1}}){str(cnt_code)}'
            AND login ~ '^[125].*'
            AND login ~ '^((?!9999).)*$'
            AND login ~ '^((?!demo).)*$';
            """

        cur.execute(fetch_query)
        if(cur.fetchone() is not None):
            df = pd.DataFrame.from_records(cur.fetchall())
            df.to_json(f"./data/db/{domain.lower()}/{domain}_{cnt_name}.json")
            print(f"JSON file created for {cnt_name}")

    con.close()
    print("Connection to DB closed")


def create_codebook(domain: str) -> pd.DataFrame:
    if domain=='SCI':
        domain_cbk = 'Sci'
    else:
        domain_cbk = domain

    path = [f for f in glob.glob("./data/*.xlsx") if "codebook_analysis" in f and domain_cbk in f]

    wb = load_workbook(filename=path[0])
    sheet_cq = [f for f in wb.sheetnames if "_CQ" in f and not 'Speaking' in f]

    if(domain == 'FLA'):

        r_cols = pd.DataFrame(
            {
                "old": [
                    'Unique item ID (Do not use - ID not in TAO)',
                    'Unit/task ID (TAO)',
                    'FLA Item ID shorter \n(11 cha MAX)',
                    'Item Format',
                    'Item type',
                    'CQ order',
                    'Key'
                ],
                "new": [
                    'qtiLabel',
                    'unit_id',
                    'qtiLabel2',
                    'item_format',
                    'item_type',
                    'item_order',
                    'key'
                ]
            }
        )

    elif(domain == 'SCI'):
        r_cols = pd.DataFrame(
            {
                "old": [
                    'CB - 2025 FT\nItem ID',
                    'CB - 2025 FT\nUnit ID',
                    'Item Format',
                    'Item type',
                    'CQ order',
                    'Key'
                ],
                "new": [
                    'qtiLabel',
                    'unit_id',
                    'item_format',
                    'item_type',
                    'item_order',
                    'key'
                ]
            }
        )

    cbk_l = []
    wb = load_workbook(filename=path[0])
    for sheet_names in sheet_cq:

        if domain == 'FLA':
            df_name = sheet_names.split("_")[1]
        else:
            df_name = domain
        print(f"Extracting codebook data from sheet {sheet_names}")
        dat = pd.read_excel(path[0],sheet_name = sheet_names)
        dat = dat.loc[dat['Stimulus']==0,]
        if domain == 'FLA':
            # dat['Unique item ID'] = dat['Unique item ID'].apply(lambda x: re.sub("FLA25","FLA",x))
            dat['domain'] = f"FLA-{df_name[0]}"
        else:
            dat['domain'] = f"{df_name}"

        dat = dat.rename(columns=lambda x: x.strip()).rename(columns=lambda x: x.replace('\xa0', '')).rename(columns=dict(zip(r_cols["old"], r_cols["new"])))
        new_cols_all = list(r_cols['new'].unique())
        new_cols_all.append('domain')
        new_cols_none = set(dat.columns) - set(new_cols_all)
        new_cols = list(set(dat.columns) - set(new_cols_none))
        dat = dat.loc[:,new_cols].drop_duplicates()
        # dat = dat.loc[:,'qtiLabel':'domain']
        key_i = [i for i,x in enumerate(dat.columns) if 'key' in x]
        dat_cols = dat.columns.to_list()
        for i,x in enumerate(key_i):
            dat_cols[x] = 'cq_key' if i == 0 else 'cq_key' + str(i + 1)
        dat.columns = dat_cols

        # dat = dat.set_axis(['qtiLabel','item_format','item_type','item_order','cq_key','cq_key2','domain'],axis = 1)
        dat['qtiLabel'] = dat['qtiLabel'].str.replace('\xa0', '')

        md_text = f"./data/metadata/{domain.lower()}/**/*.csv"
        md_path = glob.glob(md_text,recursive=True)
        md_l = []
        for file in md_path:
            csv_file = pd.read_csv(file)
            csv_file.rename(
                columns = {
                    csv_file.columns[0]: "label",
                    csv_file.columns[3]: "resp_cat",
                    csv_file.columns[4]: "db_key"
                },
                inplace=True
            )
            csv_file = csv_file.dropna(subset = ['resp_cat'])

            if(len(csv_file.resp_cat)==0):
                continue

            if(csv_file['db_key'].astype(str).str.contains("gap",na=False).any()):
                dat_list = []
                for idrow,row in csv_file.iterrows():
                    if(pd.isnull(row['db_key'])):
                        r_key = 'None'
                    else:
                        r_key = row['db_key']

                    if("gap" not in r_key):
                        dat_row = pd.DataFrame(row[['label','resp_cat','db_key']]).transpose()
                    else:
                        key_list = [re.sub("\\sgap_[0-9]","",str(x)) for x in row['db_key'].split("|")]
                        row_list = []
                        for idx,x in enumerate(key_list):
                            row_df = pd.DataFrame(row).transpose()
                            resp_cols = [col for col in row_df if col.startswith('choice_identifier')]
                            dat_r = pd.DataFrame(
                                {
                                    'label': [row['label']],
                                    'resp_cat': ['gap_' + str(idx+1)],
                                    'db_key': [x]
                                }
                            )
                            # dat_r['resp_cat'] = dat_r['resp_cat'].apply(lambda x: re.sub("RESPONSE_0","RESPONSE",x))
                            dat_final = pd.concat(
                                [
                                    dat_r.reset_index(drop = True),
                                    row_df.loc[:,resp_cols].reset_index(drop = True)
                                ],
                                axis = 1
                            )
                            row_list.append(dat_final)

                        dat_row = pd.concat(row_list,axis = 0)
                    
                    dat_list.append(dat_row)

                md_file = pd.concat(dat_list,axis = 0)
            else:
                md_file = csv_file

            id_cols = ['label','resp_cat','db_key']
            resp_cols = [col for col in md_file if col.startswith('choice_identifier')]
            all_cols = [*id_cols,*resp_cols]
            md_long = pd.melt(
                md_file.loc[:,all_cols],
                id_vars = id_cols,
                value_vars = resp_cols,
                var_name = "col",
                value_name = "val"
            ).sort_values(
                ['label','resp_cat','col']
            )

            md_long['col'] = md_long['col'].apply(
                lambda x: str(re.sub("choice_identifier_","",x))
            ).apply(
                lambda x: re.sub("7","A",x)
            ).apply(
                lambda x: re.sub("8","B",x)
            )

            md_wide = pd.pivot(
                md_long,
                index = ['label','resp_cat','db_key'],
                columns = "col",
                values = "val"
            ).reset_index()

            md_l.append(md_wide)
            
        md_resp = pd.concat(md_l).dropna(subset=['label','resp_cat'])
        md_resp['subtask'] = md_resp['resp_cat'].apply(lambda x: "RESPONSE_0" if x=="RESPONSE" else x).apply(lambda x: re.sub("RESPONSE_","",x))
        md_resp = md_resp.sort_values(['label','subtask'])
        md_resp['subtask'] = md_resp.groupby('label').cumcount() + 1
        md_resp['subtask'] = md_resp['subtask'].apply(lambda x: str(x).zfill(2))
        if (domain == 'FLA') & (df_name[0] != 'S'):
            md_resp['qtiLabel'] = np.where(md_resp['db_key'].str.contains('gap',na = False),md_resp['label'],md_resp['label'] + md_resp['subtask'])
        else:
            md_resp['qtiLabel'] = md_resp['label']
        dat = pd.merge(dat,md_resp.drop(columns = ['label']).drop_duplicates(),how = 'left',on = ['qtiLabel'])
        cbk_l.append(dat)

    cbk = pd.concat(cbk_l,ignore_index = True).drop_duplicates()
    cbk.to_excel(f'./data/codebook_{domain}.xlsx')
    print(f"Codebook created for {domain}")

    return cbk

